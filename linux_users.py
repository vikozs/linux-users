#!/usr/bin/env python3
"""
linux_users.py — account and access lifecycle audit and remediation for RHEL 9.

Sixth tool in the family (linux-audit, linux-harden, linux-diskspace,
linux-patch, linux-certs). Reuses ssh_exec.py and xlsx_safe.py.

Two modes:

    discover   Enumerate local accounts (passwd + shadow aging + last login),
               sudoers, and per-user authorized_keys. Flag duplicate UID 0,
               empty passwords, never-expiring passwords, and stale accounts.
               Write a plan (JSON) and a report (xlsx). Changes nothing.

    apply      Lock (and optionally expire) the stale accounts a plan proposes,
               RE-VALIDATING against live last-login first so an account that
               logged in since discover drops out. Never deletes an account.

Safety
------
  * Protected accounts are never touched: root, every UID < 1000, the account
    you connect as, and anything named with --protect.
  * apply only ever locks or expires. It never runs userdel, and it never
    removes SSH keys or edits sudoers. Those are report-only in this release.
  * An account whose last login cannot be determined is never auto-locked.

Run artifacts (plan JSON, xlsx) list your accounts. Keep them out of version
control (see .gitignore). Shadow hashes are never collected or written.
"""

import argparse
import datetime as _dt
import json
import re
import sys

from ssh_exec import SSHConfig, parse_hosts, run_fleet, run_one

__version__ = "1.0.0"
BUILD = "2026-07-31.initial"

PLAN_SCHEMA = "linux-users.account-plan"
RESULT_SCHEMA = "linux-users.apply-result"
SCHEMA_VERSION = "1.0"

STALE_DAYS_DEFAULT = 90
LOGIN_UID_MIN = 1000
NOLOGIN_SHELLS = {"/usr/sbin/nologin", "/sbin/nologin", "/bin/false",
                  "/usr/bin/false", "", "/bin/sync"}
WEAK_KEY_TYPES = {"ssh-dss"}          # DSA is deprecated/disabled in RHEL 9


# ---------------------------------------------------------------------------
# Remote collector
# ---------------------------------------------------------------------------

DISCOVER_SCRIPT = r"""
set -u
now=$(date +%s)
echo "===FACTS==="
echo "hostname=$(hostname -f 2>/dev/null || hostname)"
echo "distro=$( (. /etc/os-release 2>/dev/null && echo "$PRETTY_NAME") || echo unknown)"
echo "now_epoch=$now"

echo "===PASSWD==="
getent passwd

echo "===SHADOW==="
# name|status|max_days|last_change_days ; NEVER emit the hash
getent shadow 2>/dev/null | while IFS=: read -r name pw lastchg mn mx rest; do
  case "$pw" in
    "")        st=empty ;;
    "!"*|"*"*) st=locked ;;
    *)         st=set ;;
  esac
  echo "$name|$st|${mx:-}|${lastchg:-}"
done

echo "===LASTLOG==="
# user|never | user|unknown | user|<days since last login>
lastlog 2>/dev/null | tail -n +2 | while IFS= read -r line; do
  u=$(printf '%s' "$line" | awk '{print $1}')
  [ -z "$u" ] && continue
  case "$line" in
    *"Never logged in"*) echo "$u|never" ;;
    *)
      d=$(printf '%s' "$line" | awk '{print $(NF-5),$(NF-4),$(NF-3),$(NF-2),$(NF-1),$NF}')
      ep=$(date -d "$d" +%s 2>/dev/null || echo "")
      if [ -n "$ep" ]; then echo "$u|$(( (now - ep) / 86400 ))"; else echo "$u|unknown"; fi
      ;;
  esac
done

echo "===SUDOERS==="
for f in /etc/sudoers $(ls /etc/sudoers.d/* 2>/dev/null); do
  [ -f "$f" ] || continue
  grep -vE '^\s*(#|$|Defaults|@includedir|#includedir)' "$f" 2>/dev/null | while IFS= read -r e; do
    [ -n "$e" ] && echo "$f|$e"
  done
done

echo "===WHEEL==="
getent group wheel sudo 2>/dev/null | awk -F: '{print $4}' | tr ',' '\n' | sed '/^$/d' | sort -u

echo "===KEYS==="
getent passwd | while IFS=: read -r name pw uid gid gecos home shell; do
  { [ "$uid" -ge 1000 ] 2>/dev/null || [ "$uid" = 0 ]; } || continue
  for kf in "$home/.ssh/authorized_keys" "$home/.ssh/authorized_keys2"; do
    [ -f "$kf" ] || continue
    while IFS= read -r kl; do
      case "$kl" in ""|\#*) continue ;; esac
      typ=$(printf '%s' "$kl" | awk '{for(i=1;i<=NF;i++) if($i ~ /^(ssh-|ecdsa-|sk-)/){print $i; exit}}')
      cmt=$(printf '%s' "$kl" | awk '{print $NF}')
      [ -n "$typ" ] && echo "$name|$typ|$cmt"
    done < "$kf"
  done
done
echo "===END==="
"""


def build_lock_script(users, expire):
    lines = ["set -u", 'echo "===LOCK==="']
    for u in users:
        q = _shquote(u)
        lines.append("if usermod -L %s 2>/dev/null; then st=ok; else st=fail; fi" % q)
        if expire:
            lines.append("chage -E 1 %s >/dev/null 2>&1 || true" % q)
        lines.append('echo "%s|$st"' % u)
    lines.append('echo "===END==="')
    return "\n".join(lines) + "\n"


def _shquote(s):
    if re.fullmatch(r"[A-Za-z0-9._-]+", s or ""):
        return s
    return "'" + str(s).replace("'", "'\\''") + "'"


# ---------------------------------------------------------------------------
# Parsers (pure)
# ---------------------------------------------------------------------------

def split_sections(stdout):
    out, cur = {}, None
    for line in stdout.splitlines():
        s = line.strip()
        m = re.match(r"^===([A-Z]+)===$", s)
        if m:
            cur = None if m.group(1) == "END" else m.group(1)
            if cur:
                out.setdefault(cur, [])
            continue
        if cur is not None:
            out[cur].append(line)
    return out


def parse_facts(lines):
    d = {}
    for line in lines:
        if "=" in line:
            k, v = line.split("=", 1)
            d[k.strip()] = v.strip()
    return d


def parse_passwd(lines):
    out = {}
    for line in lines:
        parts = line.split(":")
        if len(parts) < 7:
            continue
        name, _pw, uid, gid, gecos, home, shell = parts[:7]
        try:
            uid_i, gid_i = int(uid), int(gid)
        except ValueError:
            continue
        out[name] = {"name": name, "uid": uid_i, "gid": gid_i, "gecos": gecos,
                     "home": home, "shell": shell}
    return out


def parse_shadow(lines):
    out = {}
    for line in lines:
        parts = line.split("|")
        if len(parts) != 4:
            continue
        name, status, mx, lastchg = parts
        out[name] = {"pw_status": status,
                     "max_days": _int(mx),
                     "last_change_days": _int(lastchg)}
    return out


def parse_lastlog(lines):
    out = {}
    for line in lines:
        parts = line.split("|")
        if len(parts) != 2:
            continue
        user, val = parts
        if val == "never":
            out[user] = {"never": True, "days": None, "known": True}
        elif val == "unknown":
            out[user] = {"never": False, "days": None, "known": False}
        else:
            d = _int(val)
            out[user] = {"never": False, "days": d, "known": d is not None}
    return out


def parse_sudoers(lines):
    out = []
    for line in lines:
        if "|" not in line:
            continue
        src, entry = line.split("|", 1)
        out.append({"source": src, "entry": entry.strip(),
                    "nopasswd": "NOPASSWD" in entry.upper()})
    return out


def parse_keys(lines):
    out = []
    for line in lines:
        parts = line.split("|")
        if len(parts) != 3:
            continue
        user, typ, comment = parts
        out.append({"user": user, "type": typ, "comment": comment,
                    "weak": typ in WEAK_KEY_TYPES})
    return out


def _int(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


# ---------------------------------------------------------------------------
# Account model + checks (pure)
# ---------------------------------------------------------------------------

def is_login_shell(shell):
    return shell not in NOLOGIN_SHELLS


def is_protected(name, uid, service_user, protect):
    return (name == "root" or uid == 0 or uid < LOGIN_UID_MIN
            or name == service_user or name in protect)


def account_flags(acc):
    flags = []
    if acc["pw_status"] == "empty":
        flags.append("empty password")
    if (acc["max_days"] in (None, 99999) or (acc["max_days"] or 0) < 0) \
            and acc["is_login"]:
        flags.append("password never expires")
    if acc["is_login"] and acc["pw_status"] == "locked":
        flags.append("login shell but password locked")
    if acc["last_login"]["never"] and acc["is_login"] and acc["uid"] >= LOGIN_UID_MIN:
        flags.append("never logged in")
    return flags


def build_accounts(sections, service_user, protect, stale_days):
    passwd = parse_passwd(sections.get("PASSWD", []))
    shadow = parse_shadow(sections.get("SHADOW", []))
    lastlog = parse_lastlog(sections.get("LASTLOG", []))
    wheel = set(l.strip() for l in sections.get("WHEEL", []) if l.strip())
    sudoers = parse_sudoers(sections.get("SUDOERS", []))
    sudo_named = {re.split(r"\s+", s["entry"])[0] for s in sudoers
                  if s["entry"] and not s["entry"].startswith("%")}

    accounts = []
    for name, p in passwd.items():
        sh = shadow.get(name, {})
        ll = lastlog.get(name, {"never": False, "days": None, "known": False})
        acc = {
            "name": name, "uid": p["uid"], "gid": p["gid"],
            "home": p["home"], "shell": p["shell"], "gecos": p["gecos"],
            "pw_status": sh.get("pw_status", "unknown"),
            "max_days": sh.get("max_days"),
            "last_change_days": sh.get("last_change_days"),
            "last_login": ll,
            "is_login": is_login_shell(p["shell"]),
            "sudo_capable": name in wheel or name in sudo_named,
        }
        acc["protected"] = is_protected(name, p["uid"], service_user, protect)
        acc["flags"] = account_flags(acc)
        acc["stale"] = bool(
            acc["is_login"] and acc["uid"] >= LOGIN_UID_MIN
            and not acc["protected"]
            and ll.get("known") and ll.get("days") is not None
            and ll["days"] > stale_days)
        accounts.append(acc)
    accounts.sort(key=lambda a: a["uid"])
    return accounts, sudoers


def fleet_checks(accounts):
    """Return dict of cross-account issues."""
    from collections import defaultdict
    by_uid = defaultdict(list)
    for a in accounts:
        by_uid[a["uid"]].append(a["name"])
    dup_uid0 = sorted(by_uid.get(0, []))
    dup_uids = {uid: names for uid, names in by_uid.items()
                if uid != 0 and len(names) > 1}
    empties = [a["name"] for a in accounts if a["pw_status"] == "empty"]
    return {
        "duplicate_uid0": dup_uid0 if len(dup_uid0) > 1 else [],
        "duplicate_uids": {str(k): sorted(v) for k, v in dup_uids.items()},
        "empty_passwords": empties,
    }


# ---------------------------------------------------------------------------
# Host records / plan
# ---------------------------------------------------------------------------

def host_record(res, service_user, protect, stale_days):
    if not res.ok:
        return {"host": res.host, "reachable": False, "error": res.error,
                "facts": {}, "accounts": [], "sudoers": [], "keys": [],
                "checks": {}, "candidates": [],
                "counts": {"accounts": 0, "stale": 0, "issues": 0}}
    sec = split_sections(res.stdout)
    facts = parse_facts(sec.get("FACTS", []))
    accounts, sudoers = build_accounts(sec, service_user, protect, stale_days)
    keys = parse_keys(sec.get("KEYS", []))
    checks = fleet_checks(accounts)
    candidates = [{"user": a["name"], "uid": a["uid"],
                   "last_login_days": a["last_login"]["days"],
                   "action": "lock_account"}
                  for a in accounts if a["stale"]]
    issues = (len(checks["duplicate_uid0"]) + len(checks["duplicate_uids"])
              + len(checks["empty_passwords"])
              + sum(1 for a in accounts if a["flags"]))
    return {
        "host": res.host,
        "hostname": facts.get("hostname", res.host),
        "reachable": True, "error": None, "facts": facts,
        "accounts": accounts, "sudoers": sudoers, "keys": keys,
        "checks": checks, "candidates": candidates,
        "counts": {"accounts": len(accounts), "stale": len(candidates),
                   "issues": issues,
                   "weak_keys": sum(1 for k in keys if k["weak"])},
    }


def _now():
    return _dt.datetime.now(_dt.timezone.utc).replace(microsecond=0).isoformat()


def build_plan(records, stale_days, service_user, protect):
    ok = [r for r in records if r["reachable"]]
    return {
        "schema": PLAN_SCHEMA,
        "schema_version": SCHEMA_VERSION,
        "generated": _now(),
        "generator": {"tool": "linux-users", "version": __version__, "build": BUILD},
        "options": {"stale_days": stale_days, "service_user": service_user,
                    "protect": sorted(protect)},
        "summary": {
            "hosts_total": len(records),
            "hosts_reachable": len(ok),
            "hosts_failed": len(records) - len(ok),
            "accounts_total": sum(r["counts"]["accounts"] for r in ok),
            "stale_candidates": sum(r["counts"]["stale"] for r in ok),
            "issues": sum(r["counts"]["issues"] for r in ok),
            "weak_keys": sum(r["counts"].get("weak_keys", 0) for r in ok),
        },
        "hosts": records,
    }


def write_plan(path, plan):
    with open(path, "w") as fh:
        json.dump(plan, fh, indent=2)
        fh.write("\n")


def load_plan(path):
    try:
        with open(path) as fh:
            plan = json.load(fh)
    except json.JSONDecodeError as e:
        raise ValueError("%s is not valid JSON: %s" % (path, e))
    if not isinstance(plan, dict) or plan.get("schema") != PLAN_SCHEMA:
        raise ValueError("not a %s file: %s" % (PLAN_SCHEMA, path))
    return plan


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------

NAVY = "1F3864"
HIGH = "C00000"
MED = "ED7D31"
LOW = "FFC000"
GOOD = "70AD47"


def write_report(path, plan):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from xlsx_safe import guard, safe_sheet_name, sweep, verify

    wb = Workbook()
    used = set()

    def sheet(title):
        return wb.create_sheet(safe_sheet_name(title, used))

    def header(ws, cols):
        ws.append(cols)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=NAVY)
            c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.freeze_panes = "A2"

    hosts = plan["hosts"]
    ok = [h for h in hosts if h["reachable"]]

    ws = wb.active
    ws.title = safe_sheet_name("Summary", used)
    header(ws, ["Host", "Distro", "Accounts", "Stale", "Issues", "Weak keys"])
    for h in sorted(ok, key=lambda x: (-x["counts"]["stale"], -x["counts"]["issues"])):
        ws.append([h["hostname"], h["facts"].get("distro", ""),
                   h["counts"]["accounts"], h["counts"]["stale"],
                   h["counts"]["issues"], h["counts"].get("weak_keys", 0)])
        for c in ws[ws.max_row]:
            guard(c)
        if h["counts"]["stale"]:
            ws.cell(ws.max_row, 4).fill = PatternFill("solid", fgColor=LOW)

    ws = sheet("Accounts")
    header(ws, ["Host", "User", "UID", "Shell", "Pw status", "Never expires",
                "Last login (days)", "Sudo", "Flags"])
    for h in ok:
        for a in h["accounts"]:
            ll = a["last_login"]
            ll_disp = ("never" if ll["never"] else
                       "unknown" if not ll["known"] else ll["days"])
            never_exp = "yes" if "password never expires" in a["flags"] else ""
            ws.append([h["hostname"], a["name"], a["uid"], a["shell"],
                       a["pw_status"], never_exp, ll_disp,
                       "yes" if a["sudo_capable"] else "",
                       "; ".join(a["flags"])])
            for c in ws[ws.max_row]:
                guard(c)
            if a["stale"]:
                ws.cell(ws.max_row, 2).fill = PatternFill("solid", fgColor=LOW)
            if "empty password" in a["flags"]:
                cell = ws.cell(ws.max_row, 5)
                cell.fill = PatternFill("solid", fgColor=HIGH)
                cell.font = Font(color="FFFFFF")

    ws = sheet("Stale Candidates")
    header(ws, ["Host", "User", "UID", "Last login (days)", "Action"])
    for h in ok:
        for c in h["candidates"]:
            ws.append([h["hostname"], c["user"], c["uid"],
                       c["last_login_days"], c["action"]])
            for cell in ws[ws.max_row]:
                guard(cell)
            ws.cell(ws.max_row, 2).fill = PatternFill("solid", fgColor=LOW)

    ws = sheet("Sudoers")
    header(ws, ["Host", "Source", "Entry", "NOPASSWD"])
    for h in ok:
        for s in h["sudoers"]:
            ws.append([h["hostname"], s["source"], s["entry"],
                       "yes" if s["nopasswd"] else ""])
            for cell in ws[ws.max_row]:
                guard(cell)
            if s["nopasswd"]:
                ws.cell(ws.max_row, 4).fill = PatternFill("solid", fgColor=MED)
                ws.cell(ws.max_row, 4).font = Font(color="FFFFFF")

    ws = sheet("SSH Keys")
    header(ws, ["Host", "User", "Key type", "Comment", "Weak"])
    for h in ok:
        for k in h["keys"]:
            ws.append([h["hostname"], k["user"], k["type"], k["comment"],
                       "yes" if k["weak"] else ""])
            for cell in ws[ws.max_row]:
                guard(cell)
            if k["weak"]:
                cell = ws.cell(ws.max_row, 5)
                cell.fill = PatternFill("solid", fgColor=HIGH)
                cell.font = Font(color="FFFFFF")

    ws = sheet("Issues")
    header(ws, ["Host", "Issue", "Detail"])
    for h in ok:
        ch = h["checks"]
        if ch.get("duplicate_uid0"):
            ws.append([h["hostname"], "duplicate UID 0",
                       ", ".join(ch["duplicate_uid0"])])
            _paint_issue(ws, PatternFill, Font, HIGH)
        for uid, names in ch.get("duplicate_uids", {}).items():
            ws.append([h["hostname"], "duplicate UID %s" % uid, ", ".join(names)])
            _paint_issue(ws, PatternFill, Font, MED)
        if ch.get("empty_passwords"):
            ws.append([h["hostname"], "empty password",
                       ", ".join(ch["empty_passwords"])])
            _paint_issue(ws, PatternFill, Font, HIGH)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            guard(c)

    ws = sheet("Errors")
    header(ws, ["Host", "Error"])
    for h in hosts:
        if not h["reachable"]:
            ws.append([h["host"], h.get("error") or "unreachable"])
            for c in ws[ws.max_row]:
                guard(c)
                c.fill = PatternFill("solid", fgColor=HIGH)
                c.font = Font(color="FFFFFF")

    ws = sheet("About")
    s = plan["summary"]
    o = plan.get("options", {})
    about = [
        ("Tool", "linux-users %s" % __version__),
        ("Build", plan["generator"].get("build", BUILD)),
        ("Generated", plan["generated"]),
        ("Stale threshold (days)", o.get("stale_days", STALE_DAYS_DEFAULT)),
        ("Service account (protected)", o.get("service_user") or ""),
        ("Extra protected", ", ".join(o.get("protect", []))),
        ("Hosts total", s["hosts_total"]),
        ("Hosts reachable", s["hosts_reachable"]),
        ("Accounts", s["accounts_total"]),
        ("Stale candidates", s["stale_candidates"]),
        ("Issues", s["issues"]),
        ("Weak keys", s["weak_keys"]),
        ("Note", "apply only locks/expires stale accounts and re-validates "
                 "last login first. It never deletes accounts, removes keys, "
                 "or edits sudoers. Protected: root, UID<1000, the service "
                 "account, and --protect names."),
    ]
    for k, v in about:
        ws.append([k, v])
        guard(ws.cell(ws.max_row, 2))
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 74

    for w in wb.worksheets:
        for col in "ABCDEFGHI":
            if w[col + "1"].value:
                w.column_dimensions[col].width = max(
                    w.column_dimensions[col].width or 0, 14)

    swept = sweep(wb)
    wb.save(path)
    bad = verify(path)
    if bad:
        raise RuntimeError("report has formula cells after sweep: %s" % bad)
    return swept


def _paint_issue(ws, PatternFill, Font, color):
    for c in ws[ws.max_row]:
        c.fill = PatternFill("solid", fgColor=color)
        c.font = Font(color="FFFFFF")


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

def say(msg):
    print(msg, file=sys.stderr, flush=True)


def do_discover(hosts, cfg, args):
    protect = set(args.protect or [])
    say("linux-users %s [build %s] — discover, %d host(s), stale>%dd"
        % (__version__, BUILD, len(hosts), args.stale_days))
    records = []
    for res in run_fleet(hosts, DISCOVER_SCRIPT, cfg, workers=args.workers):
        rec = host_record(res, args.user, protect, args.stale_days)
        records.append(rec)
        if rec["reachable"]:
            say("  %-40s %3d accounts  %d stale  %d issues"
                % (rec["hostname"], rec["counts"]["accounts"],
                   rec["counts"]["stale"], rec["counts"]["issues"]))
        else:
            say("  %-40s FAILED: %s" % (res.host, rec["error"]))
    records.sort(key=lambda r: (not r["reachable"], -r["counts"]["stale"]))
    plan = build_plan(records, args.stale_days, args.user, protect)
    write_plan(args.plan, plan)
    write_report(args.output, plan)
    s = plan["summary"]
    say("\nPlan:   %s" % args.plan)
    say("Report: %s" % args.output)
    say("Totals: %d accounts, %d stale candidates, %d issues, %d weak keys"
        % (s["accounts_total"], s["stale_candidates"], s["issues"],
           s["weak_keys"]))
    return plan


def _confirm(prompt, force):
    if force:
        return True
    try:
        return input(prompt + " [y/N] ").strip().lower() in ("y", "yes")
    except EOFError:
        return False


def do_apply(plan, hosts, cfg, args):
    protect = set(plan.get("options", {}).get("protect", [])) | set(args.protect or [])
    service_user = plan.get("options", {}).get("service_user") or args.user
    stale_days = plan.get("options", {}).get("stale_days", args.stale_days)
    say("linux-users %s [build %s] — apply%s (lock%s)"
        % (__version__, BUILD, "", ", expire" if args.expire else ""))
    by_target = {h["target"]: h for h in hosts}
    plan_by_host = {h["host"]: h for h in plan["hosts"] if h["reachable"]}
    results = []
    for host_target, prec in plan_by_host.items():
        planned = {c["user"] for c in prec["candidates"]}
        if not planned:
            continue
        host = by_target.get(host_target) or {"target": host_target,
                                              "user": None, "port": None}
        # RE-VALIDATE: re-run discover, recompute live stale set
        live = run_one(host, DISCOVER_SCRIPT, cfg)
        rec = host_record(live, service_user, protect, stale_days)
        if not rec["reachable"]:
            results.append({"host": host_target, "reachable": False,
                            "error": rec["error"], "locked": [], "skipped": []})
            say("  %-40s FAILED re-validation: %s" % (host_target, rec["error"]))
            continue
        live_stale = {c["user"] for c in rec["candidates"]}
        # extra safety: never act on a protected name even if it slipped through
        eligible = sorted(u for u in (planned & live_stale)
                          if not is_protected(
                              u, _uid_of(rec, u), service_user, protect))
        drifted = sorted(planned - live_stale)   # logged in since discover, etc.
        if not eligible:
            say("  %-40s nothing still eligible (%d drifted)"
                % (host_target, len(drifted)))
            results.append({"host": host_target, "reachable": True,
                            "locked": [], "skipped": drifted,
                            "status": "no_action"})
            continue
        prompt = ("  Lock%s %d account(s) on %s: %s?"
                  % ("/expire" if args.expire else "", len(eligible),
                     rec["hostname"], ", ".join(eligible)))
        if not _confirm(prompt, args.force):
            results.append({"host": host_target, "reachable": True,
                            "locked": [], "skipped": eligible,
                            "status": "skipped"})
            say("    skipped")
            continue
        out = run_one(host, build_lock_script(eligible, args.expire), cfg)
        locked, failed = _parse_lock(out)
        results.append({"host": host_target, "reachable": True,
                        "locked": locked, "failed": failed,
                        "skipped": drifted, "status": "applied"})
        say("    locked %d, failed %d, drifted %d"
            % (len(locked), len(failed), len(drifted)))
    _write_results(args.results, plan, results)
    _summarize(results)
    return results


def _uid_of(rec, name):
    for a in rec["accounts"]:
        if a["name"] == name:
            return a["uid"]
    return LOGIN_UID_MIN


def _parse_lock(res):
    locked, failed = [], []
    if not res.ok:
        return locked, failed
    for line in split_sections(res.stdout).get("LOCK", []):
        if "|" in line:
            u, st = line.split("|", 1)
            (locked if st.strip() == "ok" else failed).append(u)
    return locked, failed


def _write_results(path, plan, results):
    doc = {"schema": RESULT_SCHEMA, "schema_version": SCHEMA_VERSION,
           "generated": _now(),
           "generator": {"tool": "linux-users", "version": __version__, "build": BUILD},
           "plan": {"generated": plan.get("generated"),
                    "generator": plan.get("generator")},
           "hosts": results}
    with open(path, "w") as fh:
        json.dump(doc, fh, indent=2)
        fh.write("\n")
    say("Results: %s" % path)


def _summarize(results):
    locked = sum(len(r.get("locked", [])) for r in results)
    failed = sum(len(r.get("failed", [])) for r in results)
    skipped = sum(len(r.get("skipped", [])) for r in results)
    say("\nLocked: %d  failed: %d  skipped/drifted: %d" % (locked, failed, skipped))


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def build_cfg(args):
    cfg = SSHConfig(
        user=args.user, port=args.port, identity=args.identity,
        escalate=args.escalate, ask_ssh_pass=args.ask_ssh_pass,
        ssh_pass_env=args.ssh_pass_env, ask_sudo_pass=args.ask_sudo_pass,
        sudo_pass_same_as_ssh=args.sudo_pass_same_as_ssh,
        host_key_checking=args.host_key_checking, ssh_opts=args.ssh_opt,
        connect_timeout=args.connect_timeout, cmd_timeout=args.cmd_timeout)
    cfg.resolve_passwords()
    return cfg


def _add_conn_args(ap):
    ap.add_argument("-H", "--hosts", metavar="FILE", help="host list file")
    ap.add_argument("-u", "--user", help="default SSH user (also protected)")
    ap.add_argument("-p", "--port", help="default SSH port")
    ap.add_argument("-i", "--identity", help="SSH private key")
    ap.add_argument("--escalate", choices=("none", "sudo"), default="sudo")
    ap.add_argument("--ask-ssh-pass", action="store_true")
    ap.add_argument("--ssh-pass-env", metavar="VAR")
    ap.add_argument("--ask-sudo-pass", action="store_true")
    ap.add_argument("--sudo-pass-same-as-ssh", action="store_true")
    ap.add_argument("--ssh-opt", action="append", default=[], metavar="OPT")
    ap.add_argument("--host-key-checking", default="accept-new")
    ap.add_argument("--connect-timeout", type=int, default=10)
    ap.add_argument("--cmd-timeout", type=int, default=120)
    ap.add_argument("--workers", type=int, default=8)
    ap.add_argument("--protect", action="append", default=[], metavar="NAME",
                    help="account name to never touch (repeatable)")
    ap.add_argument("--stale-days", type=int, default=STALE_DAYS_DEFAULT,
                    help="last-login age that makes an account stale (default 90)")


def main(argv=None):
    ap = argparse.ArgumentParser(
        prog="linux-users",
        description="Account and access lifecycle audit and remediation (RHEL 9).")
    ap.add_argument("--version", action="version",
                    version="linux-users %s (build %s)" % (__version__, BUILD))
    sub = ap.add_subparsers(dest="cmd", required=True)

    d = sub.add_parser("discover", help="audit accounts, write plan + report")
    _add_conn_args(d)
    d.add_argument("--plan", default="users_plan.json")
    d.add_argument("-o", "--output", default="users_report.xlsx")

    a = sub.add_parser("apply", help="lock/expire stale accounts from a plan")
    _add_conn_args(a)
    a.add_argument("--plan", required=True)
    a.add_argument("--expire", action="store_true",
                   help="also set account expiry (chage -E), not just lock")
    a.add_argument("--results", default="users_results.json")
    a.add_argument("-y", "--force", action="store_true",
                   help="skip per-host confirmation")

    rp = sub.add_parser("report", help="re-render a plan to xlsx (no SSH)")
    rp.add_argument("--plan", required=True)
    rp.add_argument("-o", "--output", default="users_report.xlsx")

    args = ap.parse_args(argv)

    if args.cmd == "report":
        try:
            plan = load_plan(args.plan)
        except ValueError as e:
            say("error: %s" % e)
            return 2
        write_report(args.output, plan)
        say("Report: %s" % args.output)
        return 0

    hosts = parse_hosts(args.hosts) if getattr(args, "hosts", None) else []
    if not hosts:
        ap.error("no hosts: pass -H/--hosts FILE")
    cfg = build_cfg(args)

    if args.cmd == "discover":
        do_discover(hosts, cfg, args)
    elif args.cmd == "apply":
        try:
            plan = load_plan(args.plan)
        except ValueError as e:
            say("error: %s" % e)
            return 2
        do_apply(plan, hosts, cfg, args)
    return 0


if __name__ == "__main__":
    sys.exit(main())
