import json
import os

import pytest

import linux_users as lu
from ssh_exec import Result

FIX = os.path.join(os.path.dirname(__file__), "fixtures")
SERVICE = "local.user"


def fixture(name):
    with open(os.path.join(FIX, name)) as fh:
        return fh.read()


def rec(name="discover_host.txt", service=SERVICE, protect=None, stale=90):
    return lu.host_record(Result("web01.hostname.loc", ok=True, stdout=fixture(name)),
                          service, set(protect or []), stale)


# --- parsers ----------------------------------------------------------------

def test_split_sections():
    sec = lu.split_sections(fixture("discover_host.txt"))
    assert set(sec) >= {"FACTS", "PASSWD", "SHADOW", "LASTLOG", "SUDOERS",
                        "WHEEL", "KEYS"}


def test_parse_passwd():
    sec = lu.split_sections(fixture("discover_host.txt"))
    p = lu.parse_passwd(sec["PASSWD"])
    assert p["bob"]["uid"] == 1002
    assert p["svc-nolog"]["shell"].endswith("nologin")


def test_parse_shadow_status_and_no_hash():
    sec = lu.split_sections(fixture("discover_host.txt"))
    sh = lu.parse_shadow(sec["SHADOW"])
    assert sh["eve"]["pw_status"] == "empty"
    assert sh["root"]["pw_status"] == "locked"
    assert sh["bob"]["pw_status"] == "set"
    assert sh["bob"]["max_days"] == 99999


def test_parse_lastlog_variants():
    sec = lu.split_sections(fixture("discover_host.txt"))
    ll = lu.parse_lastlog(sec["LASTLOG"])
    assert ll["carol"]["never"] is True
    assert ll["dave"]["known"] is False           # unknown -> excluded from stale
    assert ll["bob"]["days"] == 200


def test_parse_sudoers_nopasswd_flag():
    sec = lu.split_sections(fixture("discover_host.txt"))
    su = lu.parse_sudoers(sec["SUDOERS"])
    nopw = [s for s in su if s["nopasswd"]]
    assert nopw and "deploy" in nopw[0]["entry"]


def test_parse_keys_flags_weak():
    sec = lu.split_sections(fixture("discover_host.txt"))
    keys = lu.parse_keys(sec["KEYS"])
    weak = [k for k in keys if k["weak"]]
    assert len(weak) == 1 and weak[0]["type"] == "ssh-dss"


# --- protection / stale logic -----------------------------------------------

def test_is_protected():
    assert lu.is_protected("root", 0, "local.user", set())
    assert lu.is_protected("local.user", 1100, "local.user", set())
    assert lu.is_protected("svc", 900, "local.user", set())      # UID < 1000
    assert lu.is_protected("keepme", 1001, "local.user", {"keepme"})
    assert not lu.is_protected("bob", 1002, "local.user", set())


def test_stale_candidates_are_correct():
    r = rec()
    cands = {c["user"] for c in r["candidates"]}
    assert cands == {"bob", "eve", "olduser"}
    # protected / never / unknown / active are all excluded
    assert "local.user" not in cands       # service account
    assert "backdoor" not in cands     # uid 0
    assert "carol" not in cands        # never logged in
    assert "dave" not in cands         # unknown last login
    assert "alice" not in cands        # active (5 days)
    assert "svc-nolog" not in cands    # uid < 1000


def test_protect_flag_removes_candidate():
    r = rec(protect=["bob"])
    assert "bob" not in {c["user"] for c in r["candidates"]}


def test_stale_days_threshold():
    # raise threshold above 200 -> bob no longer stale, eve/olduser (>300) remain
    r = rec(stale=250)
    cands = {c["user"] for c in r["candidates"]}
    assert "bob" not in cands
    assert {"eve", "olduser"} <= cands


# --- fleet checks -----------------------------------------------------------

def test_fleet_checks_dup_uid0_and_empty():
    r = rec()
    ch = r["checks"]
    assert set(ch["duplicate_uid0"]) == {"root", "backdoor"}
    assert ch["empty_passwords"] == ["eve"]


def test_counts():
    r = rec()
    assert r["counts"]["stale"] == 3
    assert r["counts"]["weak_keys"] == 1


def test_unreachable_host():
    r = lu.host_record(Result("dead", ok=False, error="timeout"),
                       SERVICE, set(), 90)
    assert r["reachable"] is False and r["counts"]["accounts"] == 0


# --- apply reconciliation ---------------------------------------------------

def test_reconciliation_drops_reactivated_account():
    plan_cands = {"bob", "eve", "olduser"}
    live = rec(name="discover_host_bob_active.txt")
    live_stale = {c["user"] for c in live["candidates"]}
    eligible = plan_cands & live_stale
    assert eligible == {"eve", "olduser"}   # bob logged in since discover


def test_build_lock_script_lock_only():
    s = lu.build_lock_script(["bob", "eve"], expire=False)
    assert "usermod -L bob" in s and "usermod -L eve" in s
    assert "chage -E" not in s


def test_build_lock_script_with_expire():
    s = lu.build_lock_script(["bob"], expire=True)
    assert "usermod -L bob" in s and "chage -E 1 bob" in s


def test_build_lock_script_never_userdel():
    s = lu.build_lock_script(["bob", "eve", "olduser"], expire=True)
    assert "userdel" not in s


def test_build_lock_script_quotes_hostile_name():
    s = lu.build_lock_script(["a; rm -rf /"], expire=False)
    assert "'a; rm -rf /'" in s


def test_parse_lock():
    body = "===LOCK===\nbob|ok\neve|ok\nolduser|fail\n===END===\n"
    locked, failed = lu._parse_lock(Result("h", ok=True, stdout=body))
    assert locked == ["bob", "eve"] and failed == ["olduser"]


# --- plan / IO --------------------------------------------------------------

def test_build_plan_summary():
    recs = [rec(), lu.host_record(Result("dead", ok=False, error="x"),
                                  SERVICE, set(), 90)]
    plan = lu.build_plan(recs, 90, SERVICE, set())
    s = plan["summary"]
    assert s["hosts_total"] == 2 and s["hosts_reachable"] == 1
    assert s["stale_candidates"] == 3
    assert s["weak_keys"] == 1
    assert plan["schema"] == lu.PLAN_SCHEMA


def test_plan_roundtrip(tmp_path):
    plan = lu.build_plan([rec()], 90, SERVICE, set())
    p = tmp_path / "plan.json"
    lu.write_plan(str(p), plan)
    assert lu.load_plan(str(p))["summary"]["stale_candidates"] == 3


def test_load_plan_rejects_wrong_schema(tmp_path):
    p = tmp_path / "bad.json"
    p.write_text(json.dumps({"schema": "nope"}))
    with pytest.raises(ValueError):
        lu.load_plan(str(p))


def test_plan_carries_no_password_hashes():
    # defensive: the whole plan JSON must never contain a shadow hash marker
    plan = lu.build_plan([rec()], 90, SERVICE, set())
    blob = json.dumps(plan)
    assert "$6$" not in blob and "$y$" not in blob


# --- report -----------------------------------------------------------------

def test_write_report_is_formula_clean(tmp_path):
    from xlsx_safe import verify
    recs = [rec(), lu.host_record(Result("dead", ok=False, error="x"),
                                  SERVICE, set(), 90)]
    plan = lu.build_plan(recs, 90, SERVICE, set())
    out = tmp_path / "r.xlsx"
    lu.write_report(str(out), plan)
    assert verify(str(out)) == {}
    from openpyxl import load_workbook
    names = load_workbook(str(out)).sheetnames
    for s in ("Summary", "Accounts", "Stale Candidates", "Sudoers", "SSH Keys",
              "Issues", "Errors", "About"):
        assert s in names


def test_report_neutralizes_formula_injection(tmp_path):
    from xlsx_safe import verify
    body = ("===FACTS===\nhostname=h\nnow_epoch=1\n===PASSWD===\n"
            "=cmd:x:1001:1001:x:/home/x:/bin/bash\n===SHADOW===\n"
            "=cmd|set|90|1\n===LASTLOG===\n=cmd|5\n===SUDOERS===\n"
            "===WHEEL===\n===KEYS===\n===END===\n")
    plan = lu.build_plan([lu.host_record(Result("h", ok=True, stdout=body),
                                         SERVICE, set(), 90)], 90, SERVICE, set())
    out = tmp_path / "r.xlsx"
    lu.write_report(str(out), plan)
    assert verify(str(out)) == {}


def test_discover_script_valid_bash():
    import subprocess
    r = subprocess.run(["bash", "-n"], input=lu.DISCOVER_SCRIPT, text=True,
                       capture_output=True)
    assert r.returncode == 0, r.stderr
