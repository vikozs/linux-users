#!/usr/bin/env python3
"""
xlsx_safe.py — drop-in openpyxl safety helpers.

Solves one specific, expensive class of bug: openpyxl writes any string starting
with '=' as a FORMULA. Excel then fails to parse it and "repairs" the file:

    Removed Records: Formula from /xl/worksheets/sheetN.xml part

The same behaviour is a security hole when the data comes from an untrusted
source (a remote host, a user upload, a scraped page): a value like
    =cmd|'/C calc'!A1
is spreadsheet formula injection, and it fires when a human opens the report.

Usage
-----
    from xlsx_safe import guard, sweep, verify, repair

    cell = ws.cell(row=r, column=c, value=untrusted_value)
    guard(cell)                 # per-cell, at write time

    sweep(wb)                   # belt-and-braces, immediately before wb.save()
    wb.save(path)

    verify(path)                # -> {} if clean, else {sheet_xml: count}
    repair("old.xlsx")          # fix a file produced before you had the guard

CLI
---
    python3 xlsx_safe.py verify report.xlsx
    python3 xlsx_safe.py repair report.xlsx [out.xlsx]

Why both guard() and sweep(): guard() is precise but only protects the writers
you remembered to call it from. sweep() is unconditional and catches everything,
including cells written by code you did not author. Use both.

Measured behaviour (openpyxl 3.x, verified — do not trust folklore here)
-----------------------------------------------------------------------
openpyxl types a string as a formula ONLY when it starts with '=' and is longer
than one character:

    "=2+2"       -> data_type 'f'   FORMULA (breaks Excel, triggers repair)
    "=cmd|..."   -> data_type 'f'   FORMULA
    "="          -> data_type 's'   string (bare '=' is safe)
    "+cmd"       -> data_type 's'   string
    "-1234"      -> data_type 's'   string
    "@SUM(A1)"   -> data_type 's'   string

So for .xlsx the '=' prefix is the only thing that causes the repair warning,
and an explicit string type (t="s") stops Excel evaluating the cell at all.

The '+', '-', '@' prefixes are still checked by guard() on purpose: they ARE
live injection vectors in **CSV**, where there is no per-cell type metadata and
Excel infers the type at import time. If the same data can ever be routed to CSV
(export button, pandas.to_csv, a downstream job), you want those neutralised too.
Cheap insurance; no downside.
"""

import re
import sys
import zipfile

__all__ = ["guard", "sweep", "verify", "repair", "safe_sheet_name", "TRIGGERS"]

# Leading characters that Excel/Sheets/LibreOffice may evaluate.
TRIGGERS = ("=", "+", "-", "@")


def guard(cell):
    """Force a cell to string type if its text could be read as a formula.

    Call immediately after creating any cell whose value is not a trusted
    literal. Returns the cell for chaining.
    """
    v = cell.value
    if isinstance(v, str) and v[:1] in TRIGGERS:
        cell.data_type = "s"
    return cell


def sweep(wb):
    """Convert EVERY formula-typed cell in a workbook to text. Call before save().

    Returns the number of cells neutralised.
    """
    n = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    cell.data_type = "s"
                    n += 1
    return n


def verify(path):
    """Inspect a saved .xlsx for formula cells without opening Excel.

    Returns {worksheet_xml_name: formula_count} — empty dict means clean.
    Reads the raw XML, so it validates the actual artifact, not your intent.
    """
    hits = {}
    with zipfile.ZipFile(path) as z:
        for name in z.namelist():
            if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                data = z.read(name)
                c = data.count(b"<f>")
                if c:
                    hits[name] = c
    return hits


def repair(inp, out=None):
    """Rewrite an existing workbook with all formula cells converted to text.

    Preserves values, styles, freeze panes and autofilters. Returns (out, count).
    """
    from openpyxl import load_workbook
    out = out or inp.rsplit(".", 1)[0] + "_fixed.xlsx"
    wb = load_workbook(inp)
    n = sweep(wb)
    wb.save(out)
    return out, n


_ILLEGAL_SHEET = re.compile(r"[\[\]\:\*\?\/\\]")


def safe_sheet_name(name, used=None):
    """Return an Excel-legal, unique worksheet name.

    Excel rules: max 31 chars, cannot contain []:*?/\\, cannot be empty, and
    must be unique (case-insensitively). `used` is a set you own; it is updated.
    """
    used = used if used is not None else set()
    clean = _ILLEGAL_SHEET.sub("-", str(name))[:31].strip() or "sheet"
    base, i = clean, 1
    while clean.lower() in used:
        suffix = "~%d" % i
        clean = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(clean.lower())
    return clean


def _main(argv):
    if len(argv) < 3 or argv[1] not in ("verify", "repair"):
        print(__doc__.strip().split("CLI\n---\n")[1].strip())
        return 2
    cmd, path = argv[1], argv[2]
    if cmd == "verify":
        hits = verify(path)
        if hits:
            print("FORMULA CELLS FOUND:", hits)
            return 1
        print("none — file is clean")
        return 0
    out, n = repair(path, argv[3] if len(argv) > 3 else None)
    print("Neutralised %d formula cell(s). Wrote %s" % (n, out))
    return 0


if __name__ == "__main__":
    sys.exit(_main(sys.argv))
